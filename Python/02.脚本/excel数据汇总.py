import os
from openpyxl import load_workbook

# 定义函数来处理单个Excel文件
def process_excel_file(file_path, summary_sheet):
    try:
        # 加载工作簿
        workbook = load_workbook(file_path)
        sheet = workbook.active  # 打开默认的sheet
        
        # 从工作表中读取数据
        data = [
            sheet['B2'].value,
            sheet['C3'].value,
            sheet['C4'].value,
            sheet['F4'].value
        ]
        
        # 将数据追加到汇总工作表中
        summary_sheet.append(data)
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")

# 加载汇总工作簿
try:
    summary_workbook = load_workbook('d:/files/采树汇总&.xlsx')
    summary_sheet = summary_workbook.active
except Exception as e:
    print(f"Error loading summary workbook: {e}")
    exit(1)

# 设置要读取的Excel文件目录
data_directory_path = 'D:/files/采期/'

# 遍历目录中的所有文件
try:
    for file_name in os.listdir(data_directory_path):
        # 检查文件是否为Excel文件
        if file_name.lower().endswith('.xlsx'):
            # 构建完整的文件路径
            file_path = os.path.join(data_directory_path, file_name)
            # 处理单个Excel文件
            process_excel_file(file_path, summary_sheet)
except Exception as e:
    print(f"Error during file processing: {e}")

# 保存汇总工作簿
try:
    summary_workbook.save('d:/files/temp/汇总23.xlsx')
except Exception as e:
    print(f"Error saving summary workbook: {e}")

'''
这段代码实现的功能是将一个目录下所有以.xlsx结尾的Excel文件中的特定数据提取出来，并将这些数据汇总到一个新的Excel文件中。具体来说，它会执行以下步骤：

加载一个已经存在的Excel文件作为汇总工作簿，这个文件将被用来存储从其他Excel文件中提取的数据。
遍历指定目录下的所有文件。
对于每个以.xlsx结尾的文件，假设它是一个Excel文件，代码会打开这个文件并读取默认工作表中的特定单元格数据。这些特定单元格是B2、C3、C4和F4。
将从每个文件中读取的数据追加到汇总工作簿的末尾。
在处理完所有文件后，将汇总工作簿保存到指定的路径。
最终，你会得到一个包含所有指定Excel文件特定数据的汇总Excel文件。这个过程可以用来整合多个Excel文件的数据，以便于分析和报告。
'''
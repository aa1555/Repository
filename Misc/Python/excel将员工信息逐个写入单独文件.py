from openpyxl import load_workbook

# 使用load_workbook函数从指定的路径加载员工.xlsx工作簿
wb = load_workbook('D:/demo/员工.xlsx')
# 获取工作簿的活动工作表，通常是打开Excel时默认显示的工作表
sht = wb.active

# 使用load_workbook函数从指定的路径加载绩效评分表.xlsx工作簿
wb_jx = load_workbook('D:/demo/绩效评分表.xlsx')
# 获取绩效评分表工作簿的活动工作表
sht_jx = wb_jx.active

# 使用for循环遍历员工工作表中的每一行，从第2行开始（假设第1行是标题行）
for i in range(2, sht.max_row + 1):
    # 打印当前行的第4列（D列）的单元格值，通常用于检查或调试
    print(sht.cell(i, 4).value)
    
    # 将员工工作表的第1列（A列）的值写入绩效表的B2单元格
    sht_jx['B2'] = sht.cell(i, 1).value
    # 将员工工作表的第2列（B列）的值写入绩效表的B3单元格
    sht_jx['B3'] = sht.cell(i, 2).value
    # 将员工工作表的第3列（C列）的值写入绩效表的D2单元格
    sht_jx['D2'] = sht.cell(i, 3).value
    # 将员工工作表的第4列（D列）的值写入绩效表的D3单元格
    sht_jx['D3'] = sht.cell(i, 4).value
    
    # 使用save方法保存绩效评分表工作簿，保存的文件名为员工的姓名加上.xlsx扩展名
    # 文件保存在D:/demo/绩效/目录下，如果目录不存在，将抛出错误
    wb_jx.save('D:/demo/绩效/{}.xlsx'.format(sht.cell(i, 1).value))

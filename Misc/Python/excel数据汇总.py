
import os
from openpyxl import load_workbook

# 定义函数来处理单个Excel文件
def process_excel_file(file_path, summary_sheet):
    try:
        # 加载工作簿
        workbook = load_workbook(file_path)
        sheet = workbook.active  # 打开默认的sheet
        
        # 从工作表中读取数据
        data = [
            sheet['B2'].value,
            sheet['C3'].value,
            sheet['C4'].value,
            sheet['F4'].value
        ]
        
        # 将数据追加到汇总工作表中
        summary_sheet.append(data)
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")

# 加载汇总工作簿
try:
    summary_workbook = load_workbook('d:/files/采树汇总&.xlsx')
    summary_sheet = summary_workbook.active
except Exception as e:
    print(f"Error loading summary workbook: {e}")
    exit(1)

# 设置要读取的Excel文件目录
data_directory_path = 'D:/files/采期/'

# 遍历目录中的所有文件
try:
    for file_name in os.listdir(data_directory_path):
        # 检查文件是否为Excel文件
        if file_name.lower().endswith('.xlsx'):
            # 构建完整的文件路径
            file_path = os.path.join(data_directory_path, file_name)
            # 处理单个Excel文件
            process_excel_file(file_path, summary_sheet)
except Exception as e:
    print(f"Error during file processing: {e}")

# 保存汇总工作簿
try:
    summary_workbook.save('d:/files/temp/汇总23.xlsx')
except Exception as e:
    print(f"Error saving summary workbook: {e}")

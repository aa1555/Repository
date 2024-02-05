import os
from openpyxl import load_workbook

# 加载主工作簿
wb_hz = load_workbook('d:/files/报销总表.xlsx')
sht_hz = wb_hz.active  # 激活主工作表

# 遍历指定文件夹下的所有文件
for i in os.listdir('D:/files/报销'):
    # 拼接完整的文件路径
    file_path = 'D:/files/报销/' + i
    
    # 检查文件是否为Excel文件
    if file_path.endswith('.xlsx'):
        # 加载当前遍历到的Excel文件
        wb = load_workbook(file_path)
        sht = wb.active  # 激活当前工作表
        
        # 从工作表中读取特定单元格的数据
        data = [
            sht['H2'].value,
            sht['E2'].value,
            sht['B2'].value,
            sht['C4'].value  # 这里应该是 'value' 而不是 'valve'
        ]
        
        # 打印数据（可选）
        print(data)
        
        # 将读取到的数据追加到主工作表中
        sht_hz.append(data)

# 保存主工作簿
wb_hz.save('d:/files/报销总表.xlsx')
'''
从一个指定的文件夹中读取所有的Excel文件，然后将每个Excel文件中特定单元格的数据提取出来，并将这些数据追加到一个主Excel工作簿的总表中。具体来说，代码执行以下步骤：
导入必要的库：os 和 openpyxl，前者用于操作文件系统，后者用于操作Excel文件。
加载主工作簿（报销总表.xlsx），这个工作簿将用于汇总数据。
遍历指定文件夹（D:/files/报销）中的所有文件。
对于每个文件，检查文件扩展名是否为.xlsx，即确认它是否是一个Excel文件。
如果是Excel文件，则加载该文件并激活其活动工作表。
从活动工作表的特定单元格（H2, E2, B2, C4）中读取数据，并将这些数据存储在一个列表中。
将这个列表（即从单个工作表中读取的数据）追加到主工作簿的总表中。
最后，保存主工作簿，使得所有的更改都被写入文件。
完成这些步骤后，主工作簿（报销总表.xlsx）将包含来自所有其他Excel文件中特定单元格的数据。
'''
